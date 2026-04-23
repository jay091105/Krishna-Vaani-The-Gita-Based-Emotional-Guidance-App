from __future__ import annotations

import ast
import logging
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from pymongo.errors import PyMongoError

from db import ensure_indexes, get_gita_verses_collection


logger = logging.getLogger(__name__)
BASE_DIR = Path(__file__).resolve().parent
GITA_WORKBOOK_PATH = BASE_DIR / "dataset" / "Gita.xlsx"


def _normalize_key(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _safe_int(value):
    match = re.search(r"\d+", str(value or ""))
    return int(match.group(0)) if match else None


def _chapter_label(chapter_number, chapter_title="") -> str:
    if chapter_number and int(chapter_number) > 0:
        return f"Chapter {int(chapter_number)}"
    title = str(chapter_title or "").strip()
    return title or "Chapter 1"


def _parse_verse_number(value):
    text = str(value or "").strip()
    if not text:
        return None
    match = re.search(r"(?:^|[^\d])(\d+)\s*[.\-–—]\s*(\d+)", text)
    if match:
        return int(match.group(2))
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else None


def _parse_emotion_tags(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        raw_tags = list(value)
    else:
        text = str(value).strip()
        if not text:
            return []
        try:
            parsed = ast.literal_eval(text)
        except Exception:
            parsed = [piece.strip() for piece in text.split(",")]
        raw_tags = list(parsed) if isinstance(parsed, (list, tuple, set)) else [parsed]
    return [str(tag).strip().lower() for tag in raw_tags if str(tag).strip()]


def _cell_value(row, header_index, *aliases):
    for alias in aliases:
        key = header_index.get(_normalize_key(alias))
        if key is not None and key < len(row):
            return row[key]
    return None


def _read_gita_documents() -> list[dict]:
    if not GITA_WORKBOOK_PATH.exists():
        logger.warning("Gita workbook not found at %s", GITA_WORKBOOK_PATH)
        return []

    workbook = load_workbook(GITA_WORKBOOK_PATH, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    header_index = {
        _normalize_key(header): idx
        for idx, header in enumerate(headers)
        if header is not None
    }

    documents: list[dict] = []
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        chapter_number_raw = _cell_value(row, header_index, "chapter_number", "chapter_num")
        chapter_title = str(_cell_value(row, header_index, "chapter_title") or "").strip()
        chapter_verse = str(_cell_value(row, header_index, "chapter_verse", "chapter_vers") or "").strip()
        translation = str(_cell_value(row, header_index, "translation") or "").strip()
        emotion_tags = _parse_emotion_tags(_cell_value(row, header_index, "emotion_tags", "emotion_tag"))
        what_happen = str(_cell_value(row, header_index, "What happen", "what_happen", "what happened") or "").strip()
        krishna_vani = str(_cell_value(row, header_index, "KrishnaVani", "Krishna Vani", "krishna_vaani") or "").strip()
        guidance = str(_cell_value(row, header_index, "Guidance", "guidance") or "").strip()

        chapter_number = _safe_int(chapter_number_raw) or _safe_int(chapter_title)
        if chapter_number is not None and chapter_number <= 0:
            continue

        verse_number = _parse_verse_number(chapter_verse) or row_index
        now = datetime.now(timezone.utc)
        documents.append(
            {
                "chapter_number": chapter_number,
                "chapter_title": chapter_title,
                "chapter_verse": chapter_verse,
                "translation": translation,
                "emotion_tags": emotion_tags,
                "what_happen": what_happen,
                "krishna_vaani": krishna_vani,
                "guidance": guidance,
                "verse_number": verse_number,
                "chapter_label": _chapter_label(chapter_number, chapter_title),
                "created_at": now,
                "updated_at": now,
            }
        )

    return documents


def ensure_gita_verses_loaded(force_reload: bool = False) -> int:
    ensure_indexes()
    collection = get_gita_verses_collection()

    if not force_reload and collection.estimated_document_count() > 0:
        logger.info("gita_verses already populated; skipping Excel import")
        return 0

    documents = _read_gita_documents()
    if not documents:
        logger.warning("No verse documents were read from the workbook")
        return 0

    try:
        result = collection.insert_many(documents, ordered=False)
    except PyMongoError as exc:
        logger.exception("Failed to load Gita verses into MongoDB: %s", exc)
        raise

    logger.info("Inserted %s Gita verses into MongoDB", len(result.inserted_ids))
    return len(result.inserted_ids)
